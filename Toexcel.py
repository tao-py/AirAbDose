import os
import pdfplumber
import openpyxl
import pandas as pd
from pandas import DataFrame
import time
from MyLibrary import GeneralTools as GT

def get_path(filename="环境监测报告"):
    """
    获取pdf文件路径
    return: pdf文件路径
    """
    path = []
    pwd = os.getcwd()
    pdf_path = pwd + "\\" + filename
    all_items = os.listdir(pdf_path)
    for item in all_items:
        if item.endswith(".pdf"):
            path.append(pdf_path + "\\" + item)
    return path


def extract_pdf(path, save_path="output.xlsx"):
    """
    提取pdf文件中的表格数据
    """
    with pd.ExcelWriter() as writer:
      for pdf in path:
          sheetname = pdf.split("\\")[-1].split(".")[0]
          # 读取pdf文件名
          df = DataFrame()
          pdf = pdfplumber.open(pdf)
          for page in pdf.pages:
              table = page.extract_tables()
              if table:
                for row in table:
                    if row:
                      df_ = DataFrame(row)
                      df = pd.concat([df, df_])
          df.to_excel(writer, sheet_name=sheetname, index=False)


def Dataclean(input_path, output_path):
    """
    数据清洗
    """
    wb = openpyxl.load_workbook(input_path)
    wb_out = openpyxl.Workbook()
    title_out = ["index_id", "省份", "自动站名称", "年月", "小时均值最大值", "小时均值最小值", "月均值", "空气吸收剂量率单位"]
    ws_out = wb_out.active
    ws_out.title = "Data"
    for i in range(len(title_out)):
        ws_out.cell(1, i+1, title_out[i])
    wb_out.save(output_path)
    row_out = 1
    col_out = 0

    clean_ls = ["序号", "省份", "自动站名称", "月份", "空气吸收剂量率", "最大值", "最小值", "月均值"]
    sheet_names = wb.sheetnames
    start = time.perf_counter()
    counts = len(sheet_names)
    for index,sheet_name in enumerate(sheet_names):
        GT.progress_bar(index, counts, start)
        # if index >3:
        #     break
        sheet = wb[sheet_name]
        year = sheet_name.split("年")[0]
        rows = sheet.max_row
        cols = sheet.max_column
        break_flag = 0
        for i in range(1, rows+1):
            cell = sheet.cell(i, 2)
            cell_7 = sheet.cell(i, 7)

            row_out += 1
            col_out = 0
            ws_out.cell(row_out, 8, value="μGy/h")
            for j in range(1, cols+1):
                col_out += 1
                cell = sheet.cell(i, j)
                for clean in clean_ls:
                    if clean in str(cell.value):
                        break_flag = 1
                        break
                if break_flag or cell.value == 1 or cell_7.value == None:
                    # 如果"省份"列的值为1，则跳过
                    break_flag = 0
                    row_out -= 1
                    break
                if col_out == 1:
                    ws_out.cell(row_out, col_out, value=row_out-1)
                if 1<col_out<4: # 省份、自动站名称
                    if cell.value:
                        fill_value = cell.value.replace(" ", "") # 去除空格
                        fill_value = fill_value.replace("\n", "") # 去除回车符
                        ws_out.cell(row_out, col_out, value=fill_value)
                    else:
                        ws_out.cell(row_out, col_out).value = ws_out.cell(row_out-1, col_out).value
                        # 如果为空，则填充上一个单元格的值
                if col_out == 4: # 年月
                    if cell.value:
                        year_month = year + str(100+int(cell.value))[1:]
                        year_month = int(year_month)
                        ws_out.cell(row_out, col_out, value=year_month)
                    # else:
                    #     ws_out.cell(row_out, col_out).value = int(ws_out.cell(row_out+1, col_out).value) -1
                if 4<col_out<8: # 小时均值最大值、小时均值最小值、月均值
                    if cell.value:
                        if "—" in str(cell.value) or "-" in str(cell.value):
                            ws_out.cell(row_out, col_out, value="Null")
                            # 该地当月没有测量数据
                        else:
                            ws_out.cell(row_out, col_out, value=eval(cell.value))
                            # 如果为空，则填充上一个单元格的值
                    else:
                        ws_out.cell(row_out, col_out, value="Null")
                         # 该地当月没有测量数据    
    wb_out.save(output_path)                  

if __name__ == "__main__":
    save_path = "./Out/output.xlsx"
    # path = get_path("Data")
    # extract_pdf(path,save_path)
    Dataclean(save_path, "Tosql.xlsx")